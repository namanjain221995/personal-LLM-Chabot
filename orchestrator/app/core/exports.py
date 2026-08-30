"""Tabular exports (spec §8): xlsx via openpyxl (bold header + auto column
widths) and csv. Filenames are `<slug>-<timestamp>.<ext>`. Exports are capped
at 100k rows; SQL result previews at 500 rows.

Pure module: openpyxl is imported lazily inside the xlsx writers only.

`export_xlsx` writes ONE sheet, which is what a SQL result set is.
`export_workbook` writes several — a dataset export wants its statistics and
its rows side by side — and both share `_write_sheet`, so the header styling
and column sizing can only ever be defined once.
"""
from __future__ import annotations

import csv
import re
import time
from xml.sax.saxutils import escape as _xml_escape
from pathlib import Path
from typing import List, Sequence, Tuple

#: Rows sent to the browser table.
#:
#: History: 500 stopped a 225-row answer from showing in full; 2,000 covered
#: ordinary results; 10,000 (2026-08-29) covers essentially every real
#: Salesforce answer — the largest table in this warehouse holds 28,230 rows
#: and a filtered answer is far smaller.
#:
#: Why there is still a number here. `meta.data` is not just painted, it is
#: PERSISTED per message as jsonb and shipped again on every conversation
#: reload. The browser side is no longer the constraint (the table mounts only
#: the rows in view, and the SSE parser is linear, both since 2026-08-29); the
#: database and the reload are. Raise with SQL_PREVIEW_ROW_CAP if a workload
#: genuinely needs more, and use the export path (EXPORT_ROW_CAP, 100,000) for
#: anything beyond that.
PREVIEW_ROW_CAP = 10_000
EXPORT_ROW_CAP = 100_000

_SLUG_RE = re.compile(r"[^a-z0-9]+")


def slugify(text: str, max_len: int = 40, fallback: str = "export") -> str:
    slug = _SLUG_RE.sub("-", (text or "").lower()).strip("-")
    slug = slug[:max_len].strip("-")
    return slug or fallback


def timestamped_filename(slug: str, ext: str) -> str:
    stamp = time.strftime("%Y%m%d-%H%M%S")
    return f"{slugify(slug)}-{stamp}.{ext.lstrip('.')}"


#: A preview is bounded by CELLS as well as rows, because rows alone is the
#: wrong unit: a `SELECT *` on the 268-column Interview__c emits 16 KB per ROW
#: (155 MB at 10,000 rows), while a 4-column answer is trivial at 20,000.
#: Whichever limit binds first wins, so a narrow result shows in full and a
#: very wide one is cut before it can bloat the message jsonb.
#:
#: Sized from what real answers actually weigh in this deployment, measured
#: from the stored messages: 19-22 columns at ~700 bytes per row, i.e. ~35
#: bytes per cell. 120,000 cells is therefore ~4 MB in the worst realistic
#: case — 20 columns x 6,000 rows, or the full 10,000 rows at 12 columns.
#: (An earlier note here assumed 285 bytes per row from a 10-column sample;
#: real answers are 2.5x that, which is why this is bounded by cells now.)
PREVIEW_CELL_CAP = 120_000


def preview_row_limit(columns: int, cap: int = PREVIEW_ROW_CAP) -> int:
    """Rows to show given how WIDE each row is. Always at least one row."""
    if columns <= 0:
        return cap
    return max(1, min(cap, PREVIEW_CELL_CAP // columns))


def cap_rows(rows: Sequence, cap: int) -> Tuple[list, bool]:
    """Return (rows capped to `cap`, truncated flag)."""
    if cap < 0:
        raise ValueError("cap must be >= 0")
    if len(rows) > cap:
        return list(rows[:cap]), True
    return list(rows), False


def apply_export_cap(rows: Sequence, cap: int = EXPORT_ROW_CAP) -> Tuple[list, bool]:
    """Export-specific cap (100k rows by default)."""
    return cap_rows(rows, cap)


def _cell_value(value: object) -> object:
    """Coerce values openpyxl cannot store natively (dict/list/...) to str."""
    if value is None or isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


# Excel refuses a sheet name over 31 characters or containing []:*?/\\ .
_SHEET_BAD_RE = re.compile(r"[\[\]:*?/\\]")


def safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
    """A sheet title Excel will accept: legal characters, 31 chars, non-empty."""
    cleaned = _SHEET_BAD_RE.sub(" ", str(name or "")).strip()
    return (cleaned[:31].strip() or fallback)


def _write_sheet(ws, columns: Sequence[str], rows: Sequence[Sequence]) -> None:
    """Header row in bold, then the rows, then width-to-content columns."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    ws.append([str(c) for c in columns])
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([_cell_value(v) for v in row])

    # Auto column widths: longest of header/values (sampled), padded, capped.
    sample = rows[:1000]
    for idx, col in enumerate(columns):
        longest = len(str(col))
        for row in sample:
            if idx < len(row) and row[idx] is not None:
                longest = max(longest, len(str(row[idx])))
        ws.column_dimensions[get_column_letter(idx + 1)].width = min(longest + 2, 60)


def export_workbook(
    sheets: Sequence[Tuple[str, Sequence[str], Sequence[Sequence]]],
    directory: str | Path,
    slug: str,
    cap: int = EXPORT_ROW_CAP,
) -> Tuple[Path, bool]:
    """Write a multi-sheet .xlsx from (title, columns, rows) triples.

    Returns (path, truncated) — truncated if ANY sheet hit the row cap. An
    empty `sheets` still produces a valid workbook with one empty sheet,
    because a caller that found no data should get a readable file rather
    than an exception.
    """
    from openpyxl import Workbook  # lazy: keep core imports light

    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / timestamped_filename(slug, "xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # replaced by the sheets below
    truncated = False
    used: set = set()
    for index, (title, columns, rows) in enumerate(sheets):
        name = safe_sheet_name(title, fallback=f"Sheet{index + 1}")
        while name.lower() in used:  # Excel treats titles case-insensitively
            name = safe_sheet_name(f"{name[:28]}-{index + 1}", f"Sheet{index + 1}")
        used.add(name.lower())
        capped, hit = apply_export_cap(rows, cap)
        truncated = truncated or hit
        _write_sheet(wb.create_sheet(title=name), columns, capped)

    if not wb.sheetnames:
        wb.create_sheet(title="Summary")
    wb.save(path)
    return path, truncated


def export_xlsx(
    columns: Sequence[str],
    rows: Sequence[Sequence],
    directory: str | Path,
    slug: str,
    cap: int = EXPORT_ROW_CAP,
) -> Tuple[Path, bool]:
    """Write an .xlsx file with a bold header row and auto-sized columns.

    Returns (path, truncated). Rows beyond `cap` (default 100k) are dropped.
    """
    from openpyxl import Workbook  # lazy: keep core imports light

    rows, truncated = apply_export_cap(rows, cap)
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / timestamped_filename(slug, "xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    _write_sheet(ws, columns, rows)
    wb.save(path)
    return path, truncated


def export_csv(
    columns: Sequence[str],
    rows: Sequence[Sequence],
    directory: str | Path,
    slug: str,
    cap: int = EXPORT_ROW_CAP,
) -> Tuple[Path, bool]:
    """Write a .csv file. Returns (path, truncated); same 100k row cap."""
    rows, truncated = apply_export_cap(rows, cap)
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / timestamped_filename(slug, "csv")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow([str(c) for c in columns])
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])
    return path, truncated


#: XML element names cannot start with a digit and may not contain "." or most
#: punctuation, but Salesforce column keys routinely do: `expr0` is fine,
#: `Account.Name` is not, and an aliased aggregate can be anything. Sanitising
#: to a legal Name keeps the document parseable instead of emitting something
#: only a lenient reader accepts.
_XML_NAME_BAD = re.compile(r"[^A-Za-z0-9_.\-]")


def xml_element_name(key: str, fallback: str = "field") -> str:
    """A legal XML element name for an arbitrary column key."""
    name = _XML_NAME_BAD.sub("_", str(key or "").strip()).replace(".", "_")
    name = name.lstrip("-.")
    if not name:
        return fallback
    if not (name[0].isalpha() or name[0] == "_"):
        name = f"_{name}"
    if name[:3].lower() == "xml":
        name = f"_{name}"
    return name


def rows_to_xml(
    rows: Sequence[dict],
    root: str = "records",
    row_tag: str = "record",
    cap: int = EXPORT_ROW_CAP,
) -> str:
    """Serialise rows as XML, capped like every other export.

    Attribute values are quoted (an unquoted `truncated=1` is rejected by every
    conformant parser — and it would appear precisely in the truncated case the
    attribute exists to signal), and every key is sanitised into a legal
    element name.
    """
    capped, was_capped = apply_export_cap(list(rows), cap)
    out: List[str] = ['<?xml version="1.0" encoding="UTF-8"?>']
    root_tag = xml_element_name(root, "records")
    item_tag = xml_element_name(row_tag, "record")
    out.append(
        f'<{root_tag} count="{len(capped)}"'
        + (' truncated="true"' if was_capped else "")
        + ">"
    )
    for row in capped:
        out.append(f"  <{item_tag}>")
        for key, value in (row or {}).items():
            tag = xml_element_name(key)
            text = "" if value is None else _xml_escape(str(value))
            out.append(f"    <{tag}>{text}</{tag}>")
        out.append(f"  </{item_tag}>")
    out.append(f"</{root_tag}>")
    return "\n".join(out)


def export_xml(
    rows: Sequence[dict],
    directory,
    slug: str = "export",
    cap: int = EXPORT_ROW_CAP,
) -> Tuple[Path, bool]:
    """Write a .xml file next to the csv/xlsx exports. Returns (path, truncated)."""
    capped, truncated = apply_export_cap(list(rows), cap)
    path = Path(directory) / timestamped_filename(slug, "xml")
    path.write_text(rows_to_xml(capped, cap=cap), encoding="utf-8")
    return path, truncated


__all__: List[str] = [
    "PREVIEW_ROW_CAP",
    "PREVIEW_CELL_CAP",
    "preview_row_limit",
    "EXPORT_ROW_CAP",
    "slugify",
    "timestamped_filename",
    "cap_rows",
    "apply_export_cap",
    "export_xlsx",
    "export_workbook",
    "export_csv",
    "export_xml",
    "rows_to_xml",
    "xml_element_name",
    "safe_sheet_name",
]
