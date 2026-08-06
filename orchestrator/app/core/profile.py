"""Tabular profiling (Phase 4).

The model is shown a PROFILE — shape, dtypes, null rates, ranges — never the
file. Exactly TWO things in a profile are raw data, and they are the only such
content that reaches a prompt:

  * sample rows   (PROFILE_SAMPLE_ROWS, cells truncated)
  * top values    (PROFILE_TOP_VALUES, values truncated)

Both go through `clip()`. Everything else is derived statistics.

Note what is deliberately ABSENT: min/max VALUES for string columns. Those are
an arbitrary raw cell from anywhere in the file — the alphabetically first
value can be a secret buried at row 500, and truncation cannot help when the
secret is short. String columns report min/max LENGTH instead.

Counting is done with DuckDB rather than pandas so a large CSV is never loaded
into memory — the box has ~27 GB free and an upload may be 2 GB extracted.
Nothing here executes file content: no pickle, no macros, no eval.
"""
from __future__ import annotations

import json
import os
from typing import Any, Dict, List, Optional

from ..config import settings
from . import archive

TABULAR_SUFFIXES = {".csv", ".tsv", ".txt", ".parquet", ".json", ".jsonl", ".ndjson"}
EXCEL_SUFFIXES = {".xlsx"}


def clip(value: Any) -> Any:
    """Truncate any raw value before it can reach a prompt."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    cap = settings.profile_cell_chars
    return text if len(text) <= cap else text[:cap] + "…[truncated]"


_STRINGISH = ("VARCHAR", "TEXT", "STRING", "CHAR", "BLOB", "JSON", "UUID")


def _is_stringish(dtype: str) -> bool:
    return any(token in (dtype or "").upper() for token in _STRINGISH)


def _duck():
    import duckdb  # lazy: keeps import time off the request path

    con = duckdb.connect(":memory:")
    # Profiling must not reach the NETWORK. `enable_external_access=false`
    # cannot be used here — it also blocks reading the local file we were
    # asked to profile — so the network is closed off specifically:
    # no extension can be fetched or auto-loaded, and the HTTP/S3
    # filesystems are disabled outright.
    for pragma in (
        "SET autoinstall_known_extensions=false",
        "SET autoload_known_extensions=false",
        "SET disabled_filesystems='HTTPFileSystem,S3FileSystem'",
    ):
        try:
            con.execute(pragma)
        except Exception:
            pass
    return con


def _reader_sql(path: str) -> str:
    lower = path.lower()
    quoted = path.replace("'", "''")
    if lower.endswith(".parquet"):
        return f"read_parquet('{quoted}')"
    if lower.endswith((".json", ".jsonl", ".ndjson")):
        return f"read_json_auto('{quoted}')"
    return f"read_csv_auto('{quoted}', SAMPLE_SIZE=20000, IGNORE_ERRORS=true)"


def profile_tabular(path: str, *, name: Optional[str] = None) -> Dict[str, Any]:
    """Shape, per-column statistics, and a capped sample — no bulk load."""
    rel = name or os.path.basename(path)
    out: Dict[str, Any] = {
        "file": rel,
        "bytes": os.path.getsize(path) if os.path.exists(path) else 0,
        "kind": "table",
    }
    con = _duck()
    try:
        src = _reader_sql(path)
        out["rows"] = int(con.execute(f"SELECT COUNT(*) FROM {src}").fetchone()[0])
        described = con.execute(f"DESCRIBE SELECT * FROM {src}").fetchall()
        columns = [{"name": r[0], "dtype": r[1]} for r in described]
        out["columns_total"] = len(columns)
        columns = columns[: settings.profile_max_columns]
        if len(described) > len(columns):
            out["columns_truncated"] = True

        rows = out["rows"] or 0
        for col in columns:
            ident = '"' + col["name"].replace('"', '""') + '"'
            try:
                nulls, distinct = con.execute(
                    f"SELECT COUNT(*) FILTER (WHERE {ident} IS NULL), "
                    f"COUNT(DISTINCT {ident}) FROM {src}"
                ).fetchone()
                col["null_pct"] = round(100.0 * nulls / rows, 2) if rows else 0.0
                col["distinct"] = int(distinct)
                if _is_stringish(col["dtype"]):
                    # NEVER report min/max VALUES for a string column. They are
                    # an arbitrary raw cell from anywhere in the file — the
                    # alphabetically first value is whatever happens to sort
                    # first, which can be a secret buried at row 500 — and
                    # truncation cannot help when the secret is short. Length
                    # statistics carry the useful signal with no raw content.
                    lo_len, hi_len = con.execute(
                        f"SELECT MIN(LENGTH({ident})), MAX(LENGTH({ident})) FROM {src}"
                    ).fetchone()
                    col["min_length"] = int(lo_len) if lo_len is not None else None
                    col["max_length"] = int(hi_len) if hi_len is not None else None
                else:
                    lo, hi = con.execute(
                        f"SELECT MIN({ident}), MAX({ident}) FROM {src}"
                    ).fetchone()
                    col["min"], col["max"] = clip(lo), clip(hi)
                # Top values are raw data too → capped in count AND length.
                if 0 < col["distinct"] <= 50:
                    tops = con.execute(
                        f"SELECT {ident} AS v, COUNT(*) AS n FROM {src} "
                        f"WHERE {ident} IS NOT NULL GROUP BY 1 ORDER BY n DESC "
                        f"LIMIT {settings.profile_top_values}"
                    ).fetchall()
                    col["top_values"] = [
                        {"value": clip(v), "count": int(n)} for v, n in tops
                    ]
            except Exception:
                col["stats_unavailable"] = True
        out["columns"] = columns

        sample = con.execute(
            f"SELECT * FROM {src} LIMIT {settings.profile_sample_rows}"
        ).fetchall()
        names = [c["name"] for c in columns]
        out["sample_rows"] = [
            {n: clip(v) for n, v in zip(names, row[: len(names)])} for row in sample
        ]

        # Small-file full-content path (2026-08-06): at or under the row
        # threshold — with no columns cut — the ENTIRE table goes into the
        # profile so the model can compute exact aggregates, ChatGPT-style.
        # Cells still pass through clip(); the char cap catches the
        # few-rows-but-very-wide case, falling back to profile-only.
        if (
            0 < rows <= settings.profile_full_rows_max
            and not out.get("columns_truncated")
        ):
            all_rows = con.execute(f"SELECT * FROM {src}").fetchall()
            full = [
                {n: clip(v) for n, v in zip(names, row[: len(names)])}
                for row in all_rows
            ]
            if len(json.dumps(full, default=str)) <= settings.profile_full_chars:
                out["full_rows"] = full
                out["full_content"] = True
    except Exception as exc:
        out["error"] = f"could not be read as a table: {type(exc).__name__}"
    finally:
        con.close()
    return out


def profile_excel(path: str, *, name: Optional[str] = None) -> Dict[str, Any]:
    """Profile an .xlsx — AFTER the zip-container caps have been applied.

    An .xlsx is a ZIP. Handing one straight to openpyxl would walk around
    every bomb cap in core/archive.py, so the caller must have run
    `archive.check_zip_container` first; this asserts it rather than trusting.
    """
    rel = name or os.path.basename(path)
    # Belt and braces: re-run the container caps here so no future caller can
    # reach openpyxl without them (raises ArchiveError on a bomb).
    archive.check_zip_container(path, label="spreadsheet")

    from openpyxl import load_workbook

    out: Dict[str, Any] = {
        "file": rel,
        "bytes": os.path.getsize(path),
        "kind": "spreadsheet",
        "sheets": [],
    }
    # read_only streams rows; data_only avoids evaluating anything.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets[:10]:
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None) or ()
            names = [
                str(h) if h is not None else f"column_{i + 1}"
                for i, h in enumerate(header)
            ][: settings.profile_max_columns]
            sample: List[Dict[str, Any]] = []
            counted = 0
            for row in rows_iter:
                counted += 1
                if len(sample) < settings.profile_sample_rows:
                    sample.append(
                        {n: clip(v) for n, v in zip(names, row[: len(names)])}
                    )
            out["sheets"].append(
                {
                    "name": ws.title,
                    "rows": counted,
                    "columns": [{"name": n} for n in names],
                    "sample_rows": sample,
                }
            )
    finally:
        wb.close()
    return out


def profile_file(path: str, *, name: Optional[str] = None) -> Dict[str, Any]:
    """Profile one file, choosing a reader by extension AND magic bytes."""
    rel = name or os.path.basename(path)
    lower = rel.lower()
    for suffix in archive.REFUSED_SUFFIXES:
        if lower.endswith(suffix):
            return {"file": rel, "kind": "skipped", "reason": f"refused type ({suffix})"}
    if any(lower.endswith(s) for s in EXCEL_SUFFIXES):
        if not archive.is_zip_container(path):
            return {"file": rel, "kind": "skipped", "reason": "not a real .xlsx"}
        return profile_excel(path, name=rel)
    if any(lower.endswith(s) for s in TABULAR_SUFFIXES) or archive.sniff_format(
        path
    ) == "parquet":
        return profile_tabular(path, name=rel)
    return {
        "file": rel,
        "kind": "other",
        "bytes": os.path.getsize(path) if os.path.exists(path) else 0,
    }


def profile_directory(root: str) -> List[Dict[str, Any]]:
    """Profile every readable file in an extracted archive, newest caps applied."""
    profiles: List[Dict[str, Any]] = []
    for dirpath, _dirs, files in os.walk(root):
        for fname in sorted(files):
            if len(profiles) >= settings.profile_max_files:
                return profiles
            full = os.path.join(dirpath, fname)
            rel = os.path.relpath(full, root)
            try:
                profiles.append(profile_file(full, name=rel))
            except archive.ArchiveError as exc:
                profiles.append({"file": rel, "kind": "skipped", "reason": str(exc)})
            except Exception as exc:  # a bad file must not sink the upload
                profiles.append(
                    {"file": rel, "kind": "skipped", "reason": type(exc).__name__}
                )
    return profiles


def profile_json(profiles: List[Dict[str, Any]]) -> str:
    return json.dumps(profiles, ensure_ascii=False, default=str)
