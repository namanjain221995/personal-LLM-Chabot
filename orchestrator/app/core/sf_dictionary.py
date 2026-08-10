"""The org's field dictionary: what users CALL things → what the API calls them.

People ask for "interview status", "close date", "annual revenue". Salesforce
wants `Interview_Status__c`, `CloseDate`, `AnnualRevenue`. Without a mapping the
model guesses — and its guesses look right, which is the dangerous part:
`WHERE Status__c = ...` on an object whose field is `Interview_Status__c` does
not error, it just quietly finds nothing.

The dictionary is built once from an org export (Object API Name, Object Label,
Field API Name, Field Label, Field Type) and stored in the data volume, so it
survives restarts and costs nothing per request.

It is NOT injected wholesale — an org this size is far too large for a prompt.
Each question retrieves only the objects it plausibly refers to.
"""
from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ..config import settings

DICTIONARY_PATH = os.environ.get(
    "SF_DICTIONARY_PATH", "/data/sf_dictionary.json"
)

#: Objects whose fields are injected for one question. More than a handful and
#: the prompt grows faster than the accuracy does.
MAX_OBJECTS = 4
#: Above the org's widest business objects (Account 275, Interview__c 264):
#: at 60 the truncation dropped the exact field the question needed —
#: Interview_Status__c — and the model "plausibly" guessed Status__c instead.
MAX_FIELDS_PER_OBJECT = 300
#: Picklist values are only shown for question-relevant fields, so this cap is
#: about the one long enum (Status_New_Reason__c has 14), not about volume.
MAX_PICKLIST_VALUES = 15
#: Help text is authored prose and occasionally a paragraph. One line of it
#: earns its place; the rest does not.
MAX_HELP_CHARS = 160

_WORD_RE = re.compile(r"[A-Za-z][A-Za-z0-9_]+")
#: Words that match everything and therefore rank nothing.
_STOP = {
    "the", "and", "for", "with", "from", "how", "many", "what", "which", "show",
    "give", "list", "all", "our", "are", "is", "in", "of", "to", "me", "data",
    "record", "records", "salesforce", "object", "objects", "field", "fields",
    "count", "total", "last", "this", "that", "get", "please", "name", "names",
}

_cache: Optional[Dict[str, Any]] = None


def _stem(word: str) -> str:
    """Crude singularisation, applied to BOTH sides of every comparison.

    People ask about "interviews"; the object is Interview__c. Without this the
    question that most obviously names an object failed to retrieve it. Being
    wrong in a consistent way is fine here — "status" stems to "statu", and so
    does Status__c.
    """
    if len(word) > 3 and word.endswith("ies"):
        return word[:-3] + "y"
    if len(word) > 3 and word.endswith("s") and not word.endswith("ss"):
        return word[:-1]
    return word


def _tokens(text: str) -> set:
    words = {w.lower() for w in _WORD_RE.findall(text or "")} - _STOP
    return {_stem(w) for w in words}


def build_from_rows(rows) -> Dict[str, Any]:
    """rows: (object_api, object_label, field_api, field_label, field_type)."""
    objects: Dict[str, Any] = {}
    for row in rows:
        if not row or not row[0]:
            continue
        obj_api = str(row[0]).strip()
        obj_label = str(row[1] or obj_api).strip()
        entry = objects.setdefault(
            obj_api, {"api": obj_api, "label": obj_label, "fields": []}
        )
        if len(row) > 2 and row[2]:
            entry["fields"].append(
                {
                    "api": str(row[2]).strip(),
                    "label": str(row[3] or row[2]).strip(),
                    "type": str(row[4] or "").strip(),
                }
            )
    return {"objects": objects}


def save(data: Dict[str, Any], path: str = DICTIONARY_PATH) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(data), encoding="utf-8")
    global _cache
    _cache = data


def load(path: str = DICTIONARY_PATH) -> Dict[str, Any]:
    """Cached, and never fatal — a missing dictionary just means no hints."""
    global _cache
    if _cache is not None:
        return _cache
    try:
        _cache = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        _cache = {"objects": {}}
    return _cache


def available() -> bool:
    return bool(load().get("objects"))


def _score(question_tokens: set, obj: Dict[str, Any]) -> int:
    """How strongly this object matches the question.

    Object names score highest — the question naming "Interview" should pull
    Interview__c ahead of every object that merely has a Status field.
    """
    score = 0
    obj_tokens = _tokens(obj["api"]) | _tokens(obj["label"])
    score += 6 * len(question_tokens & obj_tokens)
    for f in obj["fields"]:
        if question_tokens & (_tokens(f["api"]) | _tokens(f["label"])):
            score += 1
    return score


#: Per-record plumbing generated from a base object. AccountChangeEvent
#: carries Account's name and Account's fields, so it ties with Account on
#: every score and then eats one of the four slots. Same list as the SQL
#: grounding filter in core/schema_cache.py.
_SHADOW_SUFFIXES = ("Share", "History", "Feed", "ChangeEvent", "__hd")
#: The words that mean the asker actually wants the shadow.
_SHADOW_WORDS = {"share", "history", "feed", "changeevent", "change"}


def _is_shadow(api: str) -> bool:
    return api.endswith(_SHADOW_SUFFIXES)


def relevant_objects(question: str, limit: int = MAX_OBJECTS) -> List[Dict[str, Any]]:
    data = load()
    tokens = _tokens(question)
    if not tokens:
        return []
    wants_shadow = bool(tokens & {_stem(w) for w in _SHADOW_WORDS})
    scored = [
        (s, o)
        for o in data.get("objects", {}).values()
        if (wants_shadow or not _is_shadow(o["api"])) and (s := _score(tokens, o)) > 0
    ]
    scored.sort(key=lambda pair: (-pair[0], pair[1]["api"]))
    return [o for _s, o in scored[:limit]]


def _render_field(f: Dict[str, Any], tokens: set) -> str:
    """One field, spending prompt budget only where the question earns it.

    Every field shows its name, label and type. The expensive extras —
    picklist values, help text — appear only on the fields the question
    actually touches, because "ghosted" is worthless next to 256 other
    Interview__c fields but decisive next to Interview_Outcome__c.

    `tokens` has the object's own words removed by the caller: on Interview__c
    almost every field is named Interview_something, so leaving "interview" in
    would make every field look relevant and defeat the whole point.
    """
    head = (
        f"{f['api']} = \"{f['label']}\" ({f['type']})"
        if f.get("label") and f["label"] != f["api"]
        else f"{f['api']} ({f['type']})"
    )
    ref = f.get("ref")
    if ref:
        head += " →" + "/".join(ref)

    values = f.get("values")
    # Matching the values too, not just the name, is the point of the whole
    # exercise: nobody asks for "interviews by outcome", they ask "how many
    # were ghosted" — and Ghosted is a value of Interview_Outcome__c.
    matchable = _tokens(f["api"]) | _tokens(f.get("label") or "")
    if values:
        matchable |= _tokens(" ".join(values))
    if not tokens & matchable:
        return head
    if values:
        shown = values[:MAX_PICKLIST_VALUES]
        more = "" if len(values) <= len(shown) else f", +{len(values) - len(shown)} more"
        head += " [" + " | ".join(shown) + more + "]"
    help_text = f.get("help")
    if help_text:
        head += f" — {help_text[:MAX_HELP_CHARS]}"
    return head


def hint_for(question: str) -> str:
    """A compact API-name reference for the objects this question mentions.

    Empty when nothing matches, so the prompt is unchanged for questions the
    dictionary cannot help with.
    """
    picked = relevant_objects(question)
    if not picked:
        return ""
    tokens = _tokens(question)
    blocks = []
    for obj in picked:
        distinguishing = tokens - _tokens(obj["api"]) - _tokens(obj["label"])
        fields = obj["fields"][:MAX_FIELDS_PER_OBJECT]
        listed = "; ".join(_render_field(f, distinguishing) for f in fields)
        more = "" if len(obj["fields"]) <= len(fields) else f" … +{len(obj['fields']) - len(fields)} more"
        blocks.append(f'{obj["api"]} = "{obj["label"]}"\n  {listed}{more}')
    return (
        "Salesforce API names for the objects this question mentions "
        '(API_NAME = "the label a user would say"; →Target is what a '
        "lookup points at; [a | b] are the only valid values):\n"
        + "\n".join(blocks)
        + "\nUse the API names exactly as written above. A field name that "
        "looks plausible but is wrong returns no rows instead of an error. "
        "The same is true of a picklist value that is not listed."
    )


def build_from_xlsx(path: str) -> Dict[str, Any]:
    """Read an org export workbook (openpyxl, streaming — these are large)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header
    return build_from_rows(rows)


#: Header spellings seen across the two export tools, normalised to the five
#: positions build_from_rows expects.
_CSV_ALIASES = {
    "objectapiname": 0, "objectapi": 0, "object": 0,
    "objectlabel": 1,
    "fieldapiname": 2, "fieldapi": 2, "field": 2,
    "fieldlabel": 3,
    "fieldtype": 4, "type": 4,
}


def _header_map(header) -> Optional[Dict[int, int]]:
    """Which CSV column feeds which build_from_rows position.

    Returns None when the header is unrecognisable, so the caller can fall
    back to reading positionally the way it always did.
    """
    found: Dict[int, int] = {}
    for index, cell in enumerate(header or []):
        key = re.sub(r"[^a-z]", "", str(cell or "").lower())
        position = _CSV_ALIASES.get(key)
        if position is not None and position not in found.values():
            found[index] = position
    return found if {0, 1, 2}.issubset(set(found.values())) else None


def build_from_csv(path: str) -> Dict[str, Any]:
    """Read a flat object/field export.

    Header-aware on purpose. The AI-friendly exporter emits ObjectKind as the
    third column, so reading positionally silently produced a dictionary in
    which every field was named "Standard Object" — no error, just a hint that
    taught the model wrong names.
    """
    import csv

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        header = next(reader, None)
        mapping = _header_map(header)
        if mapping is None:
            return build_from_rows(reader)
        width = max(mapping.values()) + 1

        def remap(rows):
            for row in rows:
                out = [None] * width
                for index, position in mapping.items():
                    if index < len(row):
                        out[position] = row[index]
                yield out

        return build_from_rows(remap(reader))


def build_from_org_schema_json(path: str) -> Dict[str, Any]:
    """Read the AI-friendly exporter's org-schema.json.

    Richer than any flat export: lookup targets, picklist values and the
    help text a human wrote for the field. Those are the three things the
    model cannot guess and most often gets wrong.
    """
    raw = json.loads(Path(path).read_text(encoding="utf-8-sig"))
    objects: Dict[str, Any] = {}
    for obj in raw.get("objects", []):
        api = str(obj.get("apiName") or "").strip()
        if not api:
            continue
        fields = []
        for f in obj.get("fields", []):
            name = str(f.get("name") or "").strip()
            if not name:
                continue
            entry: Dict[str, Any] = {
                "api": name,
                "label": str(f.get("label") or name).strip(),
                "type": str(f.get("type") or "").strip(),
            }
            if f.get("referenceTo"):
                entry["ref"] = [str(r) for r in f["referenceTo"]]
            if f.get("picklistValues"):
                entry["values"] = [str(v) for v in f["picklistValues"]]
            help_text = f.get("helpText") or f.get("description")
            if help_text:
                entry["help"] = " ".join(str(help_text).split())
            fields.append(entry)
        objects[api] = {
            "api": api,
            "label": str(obj.get("label") or api).strip(),
            "fields": fields,
        }
    return {"objects": objects}


#: Keys the overlay contributes. `type` and `label` are deliberately absent:
#: the base dictionary is built from the org the platform actually queries, so
#: its answer for "what type is this" wins.
_ENRICH_KEYS = ("ref", "values", "help")


def merge(
    base: Dict[str, Any], overlay: Dict[str, Any], add_new: bool = False
) -> Tuple[Dict[str, Any], Dict[str, int]]:
    """Fold an overlay's extra field detail into an existing dictionary.

    Enrich-only by default, and that default matters. The base is built from
    the org the platform queries; the overlay came from a preprod org that runs
    ahead of it. A field present only in the overlay is therefore most likely
    one production does not have yet, and teaching the model to write it would
    reintroduce exactly the silently-wrong-name failure this module exists to
    stop. Pass add_new=True only when the overlay is known to match the org.

    Returns (merged, stats).
    """
    merged = json.loads(json.dumps(base))  # never mutate the caller's cache
    objects = merged.setdefault("objects", {})
    stats = {"objects": 0, "enriched": 0, "added": 0, "skipped": 0, "new_objects": 0}

    for api, over_obj in overlay.get("objects", {}).items():
        base_obj = objects.get(api)
        if base_obj is None:
            if not add_new:
                stats["skipped"] += len(over_obj["fields"])
                continue
            objects[api] = json.loads(json.dumps(over_obj))
            stats["new_objects"] += 1
            stats["added"] += len(over_obj["fields"])
            continue

        stats["objects"] += 1
        by_api = {f["api"]: f for f in base_obj["fields"]}
        for over_field in over_obj["fields"]:
            target = by_api.get(over_field["api"])
            if target is None:
                if not add_new:
                    stats["skipped"] += 1
                    continue
                base_obj["fields"].append(json.loads(json.dumps(over_field)))
                stats["added"] += 1
                continue
            extras = {k: over_field[k] for k in _ENRICH_KEYS if over_field.get(k)}
            if extras:
                target.update(extras)
                stats["enriched"] += 1
            if not target.get("label") or target["label"] == target["api"]:
                target["label"] = over_field.get("label") or target.get("label") or target["api"]
    return merged, stats


def main(argv=None) -> int:  # pragma: no cover - thin CLI
    import argparse

    parser = argparse.ArgumentParser(
        prog="python -m app.core.sf_dictionary",
        description="Load an org export so the model uses real API names.",
    )
    parser.add_argument("path")
    parser.add_argument("--out", default=DICTIONARY_PATH)
    parser.add_argument(
        "--merge",
        action="store_true",
        help="enrich the existing dictionary instead of replacing it "
        "(use for org-schema.json, which covers only the business objects)",
    )
    parser.add_argument(
        "--add-new",
        action="store_true",
        help="with --merge, also add fields the existing dictionary lacks. "
        "Off by default: an export from a preprod org runs ahead of "
        "production and those fields do not exist to query.",
    )
    args = parser.parse_args(argv)

    lowered = args.path.lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        builder = build_from_xlsx
    elif lowered.endswith(".json"):
        builder = build_from_org_schema_json
    else:
        builder = build_from_csv
    data = builder(args.path)

    if args.merge:
        base = load(args.out)
        data, stats = merge(base, data, add_new=args.add_new)
        print(
            f"  merged into {len(base.get('objects', {}))} existing objects: "
            f"{stats['objects']} matched, {stats['enriched']} fields enriched, "
            f"{stats['added']} added, {stats['skipped']} skipped as not-in-base"
        )

    save(data, args.out)
    objects = data["objects"]
    fields = sum(len(o["fields"]) for o in objects.values())
    enriched = sum(
        1 for o in objects.values() for f in o["fields"]
        if any(f.get(k) for k in _ENRICH_KEYS)
    )
    print(f"  {len(objects)} objects, {fields} fields ({enriched} with detail) → {args.out}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
