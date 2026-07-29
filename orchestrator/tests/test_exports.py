"""xlsx/csv exports: bold header, auto widths, re-readable, naming (spec §8)."""
import csv
import re

from openpyxl import load_workbook

from app.core.exports import export_csv, export_xlsx, slugify, timestamped_filename


def test_slugify():
    assert slugify("Pipeline Report: Q3!") == "pipeline-report-q3"
    assert slugify("   ") == "export"
    assert slugify("") == "export"


def test_timestamped_filename_pattern():
    name = timestamped_filename("My Export", "xlsx")
    assert re.fullmatch(r"my-export-\d{8}-\d{6}\.xlsx", name)


def test_export_xlsx_roundtrip(tmp_path):
    columns = ["name", "amount", "note"]
    rows = [
        ["Acme Corporation International", 1200.5, None],
        ["Globex", 99, {"nested": True}],
    ]
    path, truncated = export_xlsx(columns, rows, tmp_path, "Pipeline Report")

    assert path.exists()
    assert truncated is False
    assert re.fullmatch(r"pipeline-report-\d{8}-\d{6}\.xlsx", path.name)

    wb = load_workbook(path)
    ws = wb.active
    # Header row: values + bold font.
    assert [c.value for c in ws[1]] == columns
    assert all(c.font.bold for c in ws[1])
    # Data survives the round trip (dict coerced to str).
    assert ws.cell(row=2, column=1).value == "Acme Corporation International"
    assert ws.cell(row=2, column=2).value == 1200.5
    assert ws.cell(row=3, column=3).value == str({"nested": True})
    # Auto column widths were set; the long-name column is wider.
    width_a = ws.column_dimensions["A"].width
    width_b = ws.column_dimensions["B"].width
    assert width_a and width_b and width_a > width_b


def test_export_xlsx_cap(tmp_path):
    rows = [[i] for i in range(25)]
    path, truncated = export_xlsx(["n"], rows, tmp_path, "capped", cap=10)
    assert truncated is True
    wb = load_workbook(path)
    assert wb.active.max_row == 11  # header + 10 rows


def test_export_csv_roundtrip(tmp_path):
    columns = ["id", "value"]
    rows = [[1, "a"], [2, None]]
    path, truncated = export_csv(columns, rows, tmp_path, "csv test")
    assert truncated is False
    assert re.fullmatch(r"csv-test-\d{8}-\d{6}\.csv", path.name)
    with open(path, newline="", encoding="utf-8") as fh:
        got = list(csv.reader(fh))
    assert got == [["id", "value"], ["1", "a"], ["2", ""]]
